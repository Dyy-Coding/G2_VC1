import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def main():
    json_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(json_path, 'r') as f:
        materials = json.load(f)

    wb = Workbook()
    ws = wb.active
    headers = ['MaterialID', 'Name', 'CategoryID', 'Quantity', 'UnitPrice', 'SupplierID', 'MinStockLevel', 'ReorderLevel', 'UnitOfMeasure', 'Size', 'ImagePath', 'Description', 'CreatedAt', 'UpdatedAt', 'Brand', 'Location', 'SupplierContact', 'Status', 'WarrantyPeriod']

    ws.append(headers)

    # Header styles
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="FFFF00")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    for row_data in materials:
        ws.append([
            row_data.get('MaterialID', ''),
            row_data.get('Name', ''),
            row_data.get('CategoryID', ''),
            row_data.get('Quantity', ''),
            row_data.get('UnitPrice', ''),
            row_data.get('SupplierID', ''),
            row_data.get('MinStockLevel', ''),
            row_data.get('ReorderLevel', ''),
            row_data.get('UnitOfMeasure', ''),
            row_data.get('Size', ''),
            row_data.get('ImagePath', ''),
            row_data.get('Description', ''),
            row_data.get('CreatedAt', ''),
            row_data.get('UpdatedAt', ''),
            row_data.get('Brand', ''),
            row_data.get('Location', ''),
            row_data.get('SupplierContact', ''),
            row_data.get('Status', ''),
            row_data.get('WarrantyPeriod', ''),
        ])

    wb.save(output_path)

if __name__ == '__main__':
    main()
